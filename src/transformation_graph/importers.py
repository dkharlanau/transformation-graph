from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable
import csv
import json

import yaml

from .model import Graph, GraphValidationError


def _attributes(value: str | None, location: str) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError as exc:
        raise GraphValidationError(f"{location} attributes_json is not valid JSON: {exc}") from exc
    if not isinstance(parsed, dict):
        raise GraphValidationError(f"{location} attributes_json must contain a JSON object")
    return parsed


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _project(project_id: str, project_name: str, description: str | None = None, **extra: Any) -> dict[str, Any]:
    if not project_id or not project_name:
        raise GraphValidationError("project_id and project_name are required")
    project: dict[str, Any] = {"id": project_id, "name": project_name, **extra}
    if description:
        project["description"] = description
    return project


def _node_from_row(row: dict[str, str], location: str) -> dict[str, Any]:
    node: dict[str, Any] = {
        "id": (row.get("id") or "").strip(),
        "type": (row.get("type") or "").strip(),
        "title": (row.get("title") or "").strip(),
    }
    if (row.get("description") or "").strip():
        node["description"] = (row.get("description") or "").strip()
    if (row.get("tags") or "").strip():
        node["tags"] = [tag.strip() for tag in (row.get("tags") or "").split(";") if tag.strip()]
    attributes = _attributes(row.get("attributes_json"), location)
    if attributes:
        node["attributes"] = attributes
    return node


def _edge_from_row(row: dict[str, str], location: str) -> dict[str, Any]:
    edge: dict[str, Any] = {
        "from": (row.get("from") or "").strip(),
        "to": (row.get("to") or "").strip(),
        "type": (row.get("type") or "").strip(),
    }
    if (row.get("label") or "").strip():
        edge["label"] = (row.get("label") or "").strip()
    attributes = _attributes(row.get("attributes_json"), location)
    if attributes:
        edge["attributes"] = attributes
    return edge


def _require_columns(headers: Iterable[str], required: set[str], location: str) -> None:
    missing = required - set(headers)
    if missing:
        raise GraphValidationError(f"{location} missing required columns: {', '.join(sorted(missing))}")


def graph_from_csv(
    nodes_path: str | Path,
    edges_path: str | Path,
    project_id: str,
    project_name: str,
    description: str | None = None,
) -> Graph:
    nodes_path = Path(nodes_path)
    edges_path = Path(edges_path)
    nodes: list[dict[str, Any]] = []
    with nodes_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        _require_columns(reader.fieldnames or [], {"id", "type", "title"}, str(nodes_path))
        for line, row in enumerate(reader, start=2):
            nodes.append(_node_from_row(row, f"{nodes_path}:{line}"))

    edges: list[dict[str, Any]] = []
    with edges_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        _require_columns(reader.fieldnames or [], {"from", "to", "type"}, str(edges_path))
        for line, row in enumerate(reader, start=2):
            edges.append(_edge_from_row(row, f"{edges_path}:{line}"))

    return Graph(
        {
            "version": "0.1",
            "project": _project(project_id, project_name, description, source_format="csv"),
            "nodes": nodes,
            "edges": edges,
        }
    )


def graph_from_excel(
    workbook_path: str | Path,
    project_id: str,
    project_name: str,
    description: str | None = None,
    nodes_sheet: str = "Nodes",
    edges_sheet: str = "Edges",
) -> Graph:
    """Build a graph from a workbook using the same columns as the CSV importer."""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise GraphValidationError(
            'Excel support is optional. Install with: pip install -e ".[excel]"'
        ) from exc

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if nodes_sheet not in workbook.sheetnames:
            raise GraphValidationError(f"{workbook_path} missing worksheet '{nodes_sheet}'")
        if edges_sheet not in workbook.sheetnames:
            raise GraphValidationError(f"{workbook_path} missing worksheet '{edges_sheet}'")

        def read_rows(sheet_name: str, required: set[str]) -> list[tuple[int, dict[str, str]]]:
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            try:
                first = next(iterator)
            except StopIteration as exc:
                raise GraphValidationError(f"{workbook_path}:{sheet_name} is empty") from exc
            headers = [_text(value) for value in first]
            _require_columns(headers, required, f"{workbook_path}:{sheet_name}")
            rows: list[tuple[int, dict[str, str]]] = []
            for line, values in enumerate(iterator, start=2):
                row = {header: _text(value) for header, value in zip(headers, values) if header}
                if any(row.values()):
                    rows.append((line, row))
            return rows

        nodes = [
            _node_from_row(row, f"{workbook_path}:{nodes_sheet}:{line}")
            for line, row in read_rows(nodes_sheet, {"id", "type", "title"})
        ]
        edges = [
            _edge_from_row(row, f"{workbook_path}:{edges_sheet}:{line}")
            for line, row in read_rows(edges_sheet, {"from", "to", "type"})
        ]
    finally:
        workbook.close()

    return Graph(
        {
            "version": "0.1",
            "project": _project(
                project_id,
                project_name,
                description,
                source_format="excel",
                source_file=workbook_path.name,
            ),
            "nodes": nodes,
            "edges": edges,
        }
    )


def write_graph(graph: Graph, output_path: str | Path) -> None:
    output_path = Path(output_path)
    payload = graph.as_dict()
    if output_path.suffix.lower() in {".yaml", ".yml"}:
        output_path.write_text(
            yaml.safe_dump(payload, sort_keys=False, allow_unicode=True),
            encoding="utf-8",
        )
        return
    if output_path.suffix.lower() == ".json":
        output_path.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        return
    raise GraphValidationError("output file must end with .yaml, .yml, or .json")
